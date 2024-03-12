import openpyxl
from pdf import value_from_pdf
from datetime import datetime
# from openpyxl.styles import NamedStyle

today = datetime.now().day
month = datetime.now().strftime("%B")
month = month[0:3]
year = datetime.now().year

dict_day = {
    "TRANSPORT":"j26",
    "1% CASH HANDLING FEE": "j38",
    "Fuel Adjustment Factor":"j41",
}
dict_month = {
    "TRANSPORT":"j27",
    "1% SORTING FEE":"j30",
    "1% CASH HANDLING FEE": "j39",
    "Fuel Adjustment Factor":"j42",
    "N":"c30",
    "B":"c34"
}

def value_update(sheet,timeFrame,keyword,pdf_path,total=False):
    keyword_value = value_from_pdf(keyword,pdf_path,total)
    keyword_col = dict_day[keyword] if timeFrame=="day" else dict_month[keyword]
    sheet[keyword_col] = keyword_value


def new_excel_sheet(excel_path,pdf_path ,timeFrame):
    workbook = openpyxl.load_workbook(excel_path)

    sheet_names = workbook.sheetnames
    last_sheet_name = sheet_names[-1]
    worksheet = workbook[last_sheet_name]

    wwrksht=worksheet
    if timeFrame == "day":
        new_worksheet = workbook.create_sheet(title="DailyIntake" + str(today) + month + str(year))
        for row in worksheet.iter_rows(values_only=True):
            new_row = []
            for cell in row:
                new_row.append(cell)
            new_worksheet.append(new_row)

        # Copy styles and formatting from the original sheet to the new sheet
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                new_cell = new_worksheet[cell.coordinate]
                new_cell._style = cell._style
        wwrksht=new_worksheet

    wwrksht['b1'].value = datetime.today()
    prev_wrkdys = wwrksht['s3'].value
    wwrksht['s3'] = (prev_wrkdys + 1)

    value_update(wwrksht,timeFrame,"TRANSPORT",pdf_path)
    value_update(wwrksht,timeFrame,"1% CASH HANDLING FEE",pdf_path)
    value_update(wwrksht,timeFrame,"Fuel Adjustment Factor",pdf_path)
    if timeFrame!="day":
        value_update(wwrksht,timeFrame,"1% SORTING FEE",pdf_path)
        value_update(wwrksht,timeFrame,"N",pdf_path,total=True)
        value_update(wwrksht,timeFrame,"B",pdf_path,total=True)

    workbook.save(excel_path)

# pdf_path_month = r'C:\Users\NikhilVamsiGrandhi\OneDrive - Kanerika Software\Desktop\phoenix\pdfs\MTD_Product_Purchase_By_Date[2_Mar].pdf'
# pdf_path_day = r'C:\Users\NikhilVamsiGrandhi\OneDrive - Kanerika Software\Desktop\phoenix\pdfs\Daily_Product_Purchase_By_Date[2Mar].pdf'
# excel_path = r'C:\Users\NikhilVamsiGrandhi\OneDrive - Kanerika Software\Desktop\phoenix\excel_files\DailyIntake.xlsx'
# new_excel_sheet(excel_path,pdf_path_day,"day")
# new_excel_sheet(excel_path,pdf_path_month,"month")




